"""
用户批量导入相关的API端点
"""

import io
import logging
from typing import List
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from app.application.services.user_service import UserService
from app.models.user_models import User
from app.utils.deps import get_current_active_user, get_user_service
from app.schemas.user_schemas import UserBatchImportResponse

router = APIRouter()


@router.get("/template/download")
async def download_import_template(
    current_user: User = Depends(get_current_active_user)
):
    """下载用户批量导入Excel模板"""
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"用户 {current_user.username} 开始下载批量导入模板")
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "用户导入模板"
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        required_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        optional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 定义字段（必填和可选）
        fields = [
            # 必填字段
            ("username", "用户名*", True, "用户登录名，3-50个字符，只能包含字母、数字和下划线"),
            ("email", "电子邮箱*", True, "用户邮箱地址，必须唯一"),
            ("password", "密码*", True, "至少6个字符，需包含大写字母、小写字母、数字、特殊字符中至少3种"),
            ("full_name", "真实姓名*", True, "用户真实姓名"),
            
            # 可选字段
            ("phone", "手机号码", False, "用户手机号码"),
            ("tenant_id", "租户ID", False, "用户所属租户的ID"),
            ("organization_id", "组织ID", False, "用户所属组织的ID"),
            ("team_id", "团队ID（部门）", False, "用户所属团队（部门）的ID"),
            ("roles", "角色", False, "用户角色，多个角色用逗号分隔"),
            ("is_active", "是否激活", False, "true/false，默认为true"),
            ("is_verified", "邮箱已验证", False, "true/false，默认为false"),
            ("bio", "个人简介", False, "用户个人简介"),
            ("preferred_language", "偏好语言", False, "zh-CN/en-US，默认为zh-CN"),
            ("gender", "性别", False, "male/female/other"),
            ("country", "国家", False, "用户所在国家"),
            ("city", "城市", False, "用户所在城市"),
        ]
        
        # 写入标题行
        for col, (field_key, field_name, is_required, description) in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # 设置列宽
            ws.column_dimensions[cell.column_letter].width = 15
        
        # 写入字段说明行
        for col, (field_key, field_name, is_required, description) in enumerate(fields, 1):
            cell = ws.cell(row=2, column=col, value=description)
            cell.fill = required_fill if is_required else optional_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
        
        # 写入示例数据行
        example_data = [
            "admin001", "admin@example.com", "Admin123@", "管理员",
            "13800138000", "", "", "", "admin,user", "true", "true",
            "系统管理员", "zh-CN", "male", "中国", "北京"
        ]
        
        for col, value in enumerate(example_data, 1):
            cell = ws.cell(row=3, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 添加数据验证
        # 性别验证
        gender_validation = DataValidation(
            type="list",
            formula1='"male,female,other"',
            showErrorMessage=True,
            errorTitle="无效值",
            error="请选择：male, female, other"
        )
        ws.add_data_validation(gender_validation)
        gender_validation.add(f"N4:N1000")  # 性别列
        
        # 语言验证
        language_validation = DataValidation(
            type="list",
            formula1='"zh-CN,en-US"',
            showErrorMessage=True,
            errorTitle="无效值",
            error="请选择：zh-CN, en-US"
        )
        ws.add_data_validation(language_validation)
        language_validation.add(f"M4:M1000")  # 语言列
        
        # 布尔值验证
        bool_validation = DataValidation(
            type="list",
            formula1='"true,false"',
            showErrorMessage=True,
            errorTitle="无效值",
            error="请选择：true, false"
        )
        ws.add_data_validation(bool_validation)
        bool_validation.add(f"J4:K1000")  # 激活状态和验证状态列
        
        # 添加说明工作表
        instructions_ws = wb.create_sheet("导入说明")
        instructions = [
            "用户批量导入说明",
            "",
            "1. 必填字段说明（红色背景）：",
            "   - 用户名：3-50个字符，只能包含字母、数字和下划线，不能重复",
            "   - 电子邮箱：必须是有效的邮箱格式，不能重复", 
            "   - 密码：至少6个字符，需包含大写字母、小写字母、数字、特殊字符中至少3种",
            "   - 真实姓名：用户的真实姓名",
            "",
            "2. 可选字段说明（蓝色背景）：",
            "   - 手机号码：用户的联系电话",
            "   - 租户ID：用户所属的租户标识",
            "   - 组织ID：用户所属的组织标识",
            "   - 团队ID：用户所属的团队（部门）标识",
            "   - 角色：用户的角色，多个角色用英文逗号分隔",
            "   - 是否激活：true表示激活，false表示未激活",
            "   - 邮箱已验证：true表示已验证，false表示未验证",
            "",
            "3. 导入注意事项：",
            "   - 第1行为字段名称，第2行为字段说明，第3行为示例数据",
            "   - 实际数据请从第4行开始填写",
            "   - 用户名和邮箱必须唯一",
            "   - 密码安全要求较高，请确保符合复杂度要求",
            "   - 租户ID、组织ID、团队ID需要在系统中存在",
            "   - 角色名称需要在系统中存在",
            "",
            "4. 第三方用户体系对接：",
            "   - 支持从其他系统批量导入用户",
            "   - 确保用户名和邮箱在目标系统中唯一",
            "   - 可以预先创建租户、组织、团队和角色后再导入用户",
            "   - 建议分批导入，每批不超过1000个用户",
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = instructions_ws.cell(row=row, column=1, value=instruction)
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif instruction.startswith(("1.", "2.", "3.", "4.")):
                cell.font = Font(bold=True)
            
            # 设置列宽
            instructions_ws.column_dimensions["A"].width = 80
        
        # 将工作簿保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"用户 {current_user.username} 成功生成批量导入模板，文件大小: {output.getbuffer().nbytes} bytes")
        
        # 返回文件响应
        filename = "user_import_template.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{filename}"
            }
        )
        
    except Exception as e:
        logger.error(f"用户 {current_user.username} 下载批量导入模板失败: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"生成模板文件失败: {str(e)}"
        )


@router.post("/upload", response_model=UserBatchImportResponse)
async def upload_users_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    user_service: UserService = Depends(get_user_service),
):
    """批量导入用户Excel文件"""
    logger = logging.getLogger(__name__)
    
    logger.info(f"用户 {current_user.username} 开始批量导入用户，文件名: {file.filename}")
    
    # 验证文件格式
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        logger.warning(f"用户 {current_user.username} 上传了无效格式的文件: {file.filename}")
        raise HTTPException(status_code=400, detail="请上传Excel文件（.xlsx或.xls格式）")
    
    # 验证文件大小（限制为50MB）
    if file.size and file.size > 50 * 1024 * 1024:
        logger.warning(f"用户 {current_user.username} 上传的文件过大: {file.size} bytes")
        raise HTTPException(status_code=400, detail="文件大小不能超过50MB")
    
    try:
        contents = await file.read()
        logger.info(f"用户 {current_user.username} 文件读取完成，大小: {len(contents)} bytes")
        
        result = await user_service.batch_import_users(contents, current_user)
        
        logger.info(f"用户 {current_user.username} 批量导入完成，总计: {result.total_count}, "
                   f"成功: {result.success_count}, 失败: {result.error_count}")
        
        return result
        
    except HTTPException:
        # 重新抛出HTTP异常，保持原有状态码和消息
        raise
    except Exception as e:
        logger.error(f"用户 {current_user.username} 批量导入失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入处理失败：{str(e)}")